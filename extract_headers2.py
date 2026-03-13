import openpyxl
import json
import re
from openpyxl import load_workbook

def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s\-_/()\[\]{}:：]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text

def find_header_row(sheet, max_scan_rows=20):
    best_row = 1
    best_score = -1
    best_map = {}
    
    for row in range(1, min(sheet.max_row, max_scan_rows) + 1):
        header_map = {}
        score = 0
        for col in range(1, sheet.max_column + 1):
            raw = sheet.cell(row=row, column=col).value
            normalized = normalize_header(raw)
            if normalized:
                header_map[col] = normalized
                score += 1
        if score > best_score:
            best_score = score
            best_row = row
            best_map = header_map
    
    return best_row, best_map

wb1 = load_workbook('沃尔玛产品信息表.xlsx', data_only=True)
sheet1 = wb1['US']
source_row, source_headers_map = find_header_row(sheet1)
source_headers = {col: sheet1.cell(row=source_row, column=col).value for col in source_headers_map}

wb2 = load_workbook('filled_walmart_template.xlsx', data_only=True)
sheet2 = wb2[wb2.sheetnames[0]] if wb2.sheetnames else None
if sheet2 is None:
    raise ValueError("Template workbook has no sheets")
template_row, template_headers_map = find_header_row(sheet2)
template_headers = {col: sheet2.cell(row=template_row, column=col).value for col in template_headers_map}

result = {
    "source_headers": source_headers,
    "template_headers": template_headers
}

output_file = open('headers_extracted.json', 'w', encoding='utf-8')
json.dump(result, output_file, ensure_ascii=False, indent=2)
output_file.close()

print("Extraction complete")
