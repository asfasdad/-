import openpyxl
import json
from pathlib import Path

def find_header_row(sheet, max_scan=20):
    best_row = 1
    best_score = -1
    best_headers = []
    for row in range(1, min(sheet.max_row, max_scan) + 1):
        headers = []
        score = 0
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if val:
                headers.append(str(val))
                score += 1
            else:
                headers.append('')
        if score > best_score:
            best_score = score
            best_row = row
            best_headers = headers
    return best_row, [h for h in best_headers if h]

source_wb = openpyxl.load_workbook('沃尔玛产品信息表.xlsx', data_only=True)
source_sheet = source_wb.active
source_row, source_headers = find_header_row(source_sheet)

template_wb = openpyxl.load_workbook('filled_walmart_template.xlsx', data_only=True)
template_sheet = template_wb.active
template_row, template_headers = find_header_row(template_sheet)

result = {
    "source_headers": source_headers,
    "template_headers": template_headers,
    "source_header_row": source_row,
    "template_header_row": template_row
}

print(json.dumps(result, ensure_ascii=False, indent=2))
