from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app import app


def build_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    if not isinstance(ws, Worksheet):
        raise RuntimeError("Template active sheet is not worksheet")
    ws.cell(row=1, column=1, value="Seller SKU")
    ws.cell(row=1, column=2, value="Product Name")
    ws.cell(row=1, column=3, value="Brand")
    ws.cell(row=1, column=4, value="Price")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_product_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    if not isinstance(ws, Worksheet):
        raise RuntimeError("Product active sheet is not worksheet")
    ws.cell(row=1, column=1, value="SKU")
    ws.cell(row=1, column=2, value="商品名称")
    ws.cell(row=1, column=3, value="品牌")
    ws.cell(row=1, column=4, value="售价")
    ws.cell(row=2, column=1, value="SKU-001")
    ws.cell(row=2, column=2, value="测试产品A")
    ws.cell(row=2, column=3, value="BrandA")
    ws.cell(row=2, column=4, value=19.99)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main() -> None:
    client = TestClient(app)

    response = client.post(
        "/autofill",
        files={
            "template_file": (
                "template.xlsx",
                build_template_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "product_file": (
                "products.xlsx",
                build_product_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    print("status:", response.status_code)
    print("mapped:", response.headers.get("X-Mapped-Columns"))
    print("filled_rows:", response.headers.get("X-Filled-Rows"))

    out_wb = load_workbook(io.BytesIO(response.content), data_only=True)
    out_ws = out_wb.active
    if not isinstance(out_ws, Worksheet):
        raise RuntimeError("Output active sheet is not worksheet")
    print("row2:", [out_ws.cell(row=2, column=i).value for i in range(1, 5)])

    ai_response = client.post(
        "/autofill",
        data={"use_ai": "true"},
        files={
            "template_file": (
                "template.xlsx",
                build_template_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "product_file": (
                "products.xlsx",
                build_product_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    print("ai_status:", ai_response.status_code)
    print("ai_warning:", ai_response.headers.get("X-AI-Warning"))


if __name__ == "__main__":
    main()
