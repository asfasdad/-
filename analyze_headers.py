import json
from pathlib import Path
from openpyxl import load_workbook

def extract_headers(file_path):
    """Extract header row from an Excel file"""
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        # Find the header row (first row with content)
        for row in range(1, min(sheet.max_row, 10) + 1):
            headers = {}
            has_content = False
            for col in range(1, sheet.max_column + 1):
                value = sheet.cell(row=row, column=col).value
                if value:
                    headers[col] = str(value).strip()
                    has_content = True
            if has_content and len(headers) > 3:  # Assume header row has multiple columns
                return headers
        return {}
    except Exception as e:
        return {"error": str(e)}

# Analyze all completed files
completed_dir = Path("填写完成的表格")
results = {}

for file in completed_dir.glob("*.xlsx"):
    if file.name.startswith("~$"):
        continue
    headers = extract_headers(file)
    results[file.name] = headers

# Print results
print(json.dumps(results, ensure_ascii=False, indent=2))

# Also check product info table
product_file = Path("沃尔玛产品信息表.xlsx")
if product_file.exists():
    print("\n\n=== 沃尔玛产品信息表.xlsx ===")
    headers = extract_headers(product_file)
    print(json.dumps(headers, ensure_ascii=False, indent=2))
